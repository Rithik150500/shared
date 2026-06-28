"""Synchronous broadcast driver CLI for the ``munshi_welcome_video_v1`` template.

Sends the approved marketing template (VIDEO header) to one tier of the
broadcast list, paced and resumable. Dry-run by default.

Usage (dry-run, no network):
    python -m whatsapp_delivery.tools.broadcast_send \\
        --xlsx /path/to/broadcast.xlsx \\
        --campaign munshi_launch_2026_06 \\
        --tier T1 \\
        --dry-run

Live send requires ``--yes`` and exactly ONE of ``--video-file`` /
``--video-media-id``:
    python -m whatsapp_delivery.tools.broadcast_send \\
        --xlsx /path/to/broadcast.xlsx \\
        --campaign munshi_launch_2026_06 \\
        --tier T1 \\
        --video-media-id <meta_media_id> \\
        --yes

Session durability: a fresh ``get_session()`` context is used per recipient
so that a crash mid-run leaves already-sent rows committed to the DB. The
``claim_send`` → ``send`` → ``mark_sent`` sequence per recipient is the unit
of work; each is committed independently.

Phone numbers are redacted to last-4 digits in all log output (DPDP / PII
hygiene, matching ``whatsapp_delivery.dispatch.worker._redact_phone``).
"""
from __future__ import annotations

import argparse
import logging
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from data_access.daos import broadcast_dao as _dao
from data_access.engine import get_session
from whatsapp_delivery.config import WhatsAppConfig
from whatsapp_delivery.errors import MetaInvalidMessage, MetaTransientError
from whatsapp_delivery.meta_client import MetaClient
from whatsapp_delivery.template_client import TemplateClient
from whatsapp_delivery.tools.broadcast_report import summarize


log = logging.getLogger(__name__)

_DEFAULT_TEMPLATE = "munshi_welcome_video_v1"
_DEFAULT_LANGUAGE = "en"
_DEFAULT_PER_RUN_CAP = 100
_DEFAULT_DAILY_CAP = 200
_DEFAULT_SPACING_SECONDS = 2.0
_DEFAULT_RETRY_BACKOFF = 5.0

# Auto-pause guard defaults
_DEFAULT_MAX_FAIL_RATE = 0.1
_DEFAULT_MAX_UNDELIVERABLE = 0.4
# 131049-rate ceiling (incident 2026-06-25): an individual 131049 ("healthy
# ecosystem engagement") is a retryable per-user marketing cap, but a SUSTAINED
# HIGH RATE of it is Meta actively throttling the account for spam — the direct
# precursor to enforcement. The spam-flagged run was ~32% 131049, so 0.30 halts
# that scenario while still tolerating the occasional per-user cap.
_DEFAULT_MAX_MARKETING_CAP_RATE = 0.3
_DEFAULT_MIN_SAMPLE = 50


# ---------------------------------------------------------------------------
# Data types
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Recipient:
    """A single broadcast target, ready to send."""

    wa_digits: str
    name: str


# ---------------------------------------------------------------------------
# Pure helpers (unit-testable without DB / network)
# ---------------------------------------------------------------------------


def _redact_phone(phone: str | None) -> str:
    """Reduce a phone number to ``***<last 4 digits>`` for log lines.

    Mirrors the convention in ``whatsapp_delivery.dispatch.worker._redact_phone``.
    """
    if not phone:
        return "***"
    digits = "".join(c for c in str(phone) if c.isdigit())
    if len(digits) < 4:
        return "***"
    return f"***{digits[-4:]}"


def select_recipients(
    rows: list[dict],
    *,
    tier: str,
    suppressed: set[str],
    already_done: set[str],
    limit: int,
    name_fallback: str = "there",
) -> list[Recipient]:
    """Filter and cap rows for a single broadcast run.

    Args:
        rows: list of dicts with keys ``"WA_Digits"``, ``"Name (clean)"``,
              ``"Tier label"``.
        tier: prefix to match against ``"Tier label"`` (e.g. ``"T1"``).
        suppressed: set of ``wa_digits`` to skip (opt-out / deny list).
        already_done: set of ``wa_digits`` already in the ledger for this
              campaign (any status — skip to avoid double-send).
        limit: maximum number of ``Recipient`` objects to return.
        name_fallback: name to use when ``"Name (clean)"`` is blank.

    Returns:
        list of :class:`Recipient` (at most ``limit`` items).
    """
    results: list[Recipient] = []
    for row in rows:
        if len(results) >= limit:
            break
        wa_digits = row.get("WA_Digits")
        # Skip blank / None phone numbers
        if not wa_digits or not str(wa_digits).strip():
            continue
        # Normalize to digits-only so suppression/done matching is format-agnostic.
        # The stop-handler stores wa_digits as pure E.164 digits (no leading +),
        # so "91 99999-99999" and "+919999999999" must both match "919999999999".
        wa_digits = "".join(ch for ch in str(wa_digits) if ch.isdigit())
        if not wa_digits:
            continue
        # Tier filter: "Tier label" must start with the requested tier prefix
        tier_label = str(row.get("Tier label", "") or "")
        if not tier_label.startswith(tier):
            continue
        # Skip suppressed
        if wa_digits in suppressed:
            continue
        # Skip already in ledger
        if wa_digits in already_done:
            continue
        # Resolve name. Guard against pandas reading a blank cell as NaN
        # (a *truthy* float whose str() is the literal "nan"), so generic-greeting
        # rows fall back to the friendly default instead of "Hi nan,".
        raw_name = row.get("Name (clean)")
        name = "" if raw_name is None else str(raw_name).strip()
        if not name or name.lower() == "nan":
            name = name_fallback
        results.append(Recipient(wa_digits=wa_digits, name=name))
    return results


def _load_rows(xlsx_path: str) -> list[dict]:
    """Load the broadcast list from an xlsx file.

    Reads sheet ``"Broadcast List"`` using openpyxl (pure-Python, no pandas
    dependency) and returns ``list[dict]`` with the header row as keys.
    Every cell is coerced to ``str``; blank/None cells become ``""``.
    openpyxl reads integer-valued numeric cells as ``int`` (so a phone like
    ``916001141186`` comes back as the int ``916001141186``, which ``str()``
    turns into ``"916001141186"`` with no ``.0`` suffix — same result as
    pandas ``dtype=str``).  Kept thin so the send loop can be tested with
    synthetic row lists without touching the filesystem.
    """
    from openpyxl import load_workbook  # lazy import — openpyxl; not needed in pure-function tests

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Broadcast List"]
        it = ws.iter_rows(values_only=True)
        header = ["" if c is None else str(c) for c in next(it)]
        rows = []
        for raw in it:
            row = {}
            for i, h in enumerate(header):
                if not h:
                    continue
                v = raw[i] if i < len(raw) else None
                row[h] = "" if v is None else str(v)
            rows.append(row)
        return rows
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Send loop
# ---------------------------------------------------------------------------


def run(args: Any, *, rows: list[dict] | None = None) -> int:  # noqa: C901 (complexity acceptable for a CLI driver)
    """Execute the broadcast send loop.

    Args:
        args: parsed ``argparse.Namespace`` (or any object with the same
              attributes — useful for testing with ``SimpleNamespace``).
        rows: pre-loaded row dicts. If ``None``, ``_load_rows(args.xlsx)``
              is called. Injected by tests to avoid needing a real xlsx file.

    Returns:
        Exit code (0 = success, 1 = fatal error).
    """
    campaign: str = args.campaign
    tier: str = args.tier
    template: str = args.template
    language: str = args.language
    dry_run: bool = getattr(args, "dry_run", True)
    spacing_seconds: float = float(args.spacing_seconds)
    retry_backoff: float = float(args.retry_backoff)
    per_run_cap: int = int(args.per_run_cap)
    daily_cap: int = int(args.daily_cap)
    video_media_id: str | None = getattr(args, "video_media_id", None)
    video_file: str | None = getattr(args, "video_file", None)
    max_fail_rate: float = float(getattr(args, "max_fail_rate", _DEFAULT_MAX_FAIL_RATE))
    max_undeliverable: float = float(getattr(args, "max_undeliverable", _DEFAULT_MAX_UNDELIVERABLE))
    max_marketing_cap_rate: float = float(
        getattr(args, "max_marketing_cap_rate", _DEFAULT_MAX_MARKETING_CAP_RATE)
    )
    min_sample: int = int(getattr(args, "min_sample", _DEFAULT_MIN_SAMPLE))

    # Load rows from xlsx unless injected by test
    if rows is None:
        rows = _load_rows(args.xlsx)

    # --- Determine suppressed + already-done sets via a single session ---
    with get_session() as s:
        suppressed = _dao.load_suppressed_set(s)
        already_done = _dao.already_done_set(s, campaign)
        sent_24h = _dao.sent_count_since(s, campaign, hours=24)

        # --- Auto-pause guard (evaluated before any send or upload) ---
        # Reads the campaign's recent outcome from the ledger. If quality signals
        # are bad, refuses to start regardless of --dry-run or --yes.
        stats = summarize(s, campaign, tier=tier)
        if stats["attempted"] >= min_sample:
            # 131049 (marketing-cap) is a transient per-user cap — the recipient
            # hit Meta's cross-business daily marketing limit and is retryable
            # next day. It is NOT a quality signal, so exclude it from the
            # fail-rate; genuine quality failures (blocks/spam/undeliverable)
            # still count. Undeliverable also has its own undel_rate gate.
            non_transient_failed = stats["failed"] - stats["marketing_capped"]
            fail_rate = non_transient_failed / stats["attempted"]
            undel_rate = stats["undeliverable"] / stats["attempted"]
            # 131049-rate ceiling (incident 2026-06-25): a single 131049 is a
            # retryable per-user marketing cap and stays out of fail_rate, but a
            # SUSTAINED HIGH RATE means Meta is throttling the whole account for
            # spam — halt before that tips into a policy enforcement.
            mktcap_rate = stats["marketing_capped"] / stats["attempted"]
            if (
                fail_rate > max_fail_rate
                or undel_rate > max_undeliverable
                or mktcap_rate > max_marketing_cap_rate
            ):
                log.error(
                    "auto-pause: campaign=%s fail_rate=%.2f (excl %d marketing-cap) "
                    "undel_rate=%.2f mktcap_rate=%.2f attempted=%d — refusing to send",
                    campaign,
                    fail_rate,
                    stats["marketing_capped"],
                    undel_rate,
                    mktcap_rate,
                    stats["attempted"],
                )
                return 3

    recipients = select_recipients(
        rows,
        tier=tier,
        suppressed=suppressed,
        already_done=already_done,
        limit=per_run_cap,
    )

    # --- Daily-cap guard ---
    remaining_daily = max(0, daily_cap - sent_24h)
    if len(recipients) > remaining_daily:
        trimmed = len(recipients) - remaining_daily
        log.info(
            "daily_cap trim: campaign=%s daily_cap=%d sent_24h=%d "
            "candidates=%d trimming=%d remaining=%d",
            campaign,
            daily_cap,
            sent_24h,
            len(recipients),
            trimmed,
            remaining_daily,
        )
        recipients = recipients[:remaining_daily]

    if dry_run:
        log.info(
            "DRY RUN — campaign=%s tier=%s planned=%d (no sends, no uploads)",
            campaign,
            tier,
            len(recipients),
        )
        for r in recipients[:5]:
            log.info("  would-send to=%s name=%s", _redact_phone(r.wa_digits), r.name)
        if len(recipients) > 5:
            log.info("  ... and %d more", len(recipients) - 5)
        return 0

    # --- Live send path ---
    # Defense-in-depth: run() checks --yes independently of main() so that
    # calling run(SimpleNamespace(dry_run=False, yes=False)) from code or tests
    # never accidentally fires live sends.  main() also flips dry_run on --yes
    # but that cannot protect direct run() callers.
    if not getattr(args, "yes", False):
        log.error("refusing to send live without --yes")
        return 2

    phone_number_id = getattr(args, "phone_number_id", None)
    access_token = getattr(args, "access_token", None)
    if not phone_number_id or not access_token:
        cfg = WhatsAppConfig()
        phone_number_id = cfg.meta_phone_number_id
        access_token = cfg.meta_access_token

    # Upload video once (unless a pre-uploaded media_id was supplied)
    media_id: str
    if video_media_id:
        media_id = video_media_id
        log.info("using pre-supplied video media_id=%s", media_id)
    elif video_file:
        video_bytes = Path(video_file).read_bytes()
        log.info("uploading video file=%s (%d bytes)", video_file, len(video_bytes))
        mc = MetaClient(phone_number_id=phone_number_id, access_token=access_token)
        media_id = mc.upload_media(
            data=video_bytes,
            filename="munshi_welcome.mp4",
            mime_type="video/mp4",
        )
        log.info("upload complete media_id=%s", media_id)
    else:
        log.error("live send requires --video-file or --video-media-id")
        return 1

    client = TemplateClient(phone_number_id=phone_number_id, access_token=access_token)

    attempted = 0
    sent = 0
    failed = 0
    skipped = 0

    for r in recipients:
        # Per-recipient durability: use a fresh session so each send is
        # committed independently. A crash between sends leaves completed
        # rows persisted.
        with get_session() as s:
            claimed = _dao.claim_send(
                s,
                campaign=campaign,
                wa_digits=r.wa_digits,
                tier=tier,
                template_name=template,
                language=language,
            )
            if not claimed:
                log.debug(
                    "skip to=%s reason=already_claimed", _redact_phone(r.wa_digits)
                )
                skipped += 1
                time.sleep(spacing_seconds)
                continue

        attempted += 1
        log.info(
            "send attempt=%d to=%s name=%s template=%s",
            attempted,
            _redact_phone(r.wa_digits),
            r.name,
            template,
        )

        send_ok = False
        for attempt in range(3):
            try:
                with get_session() as s:
                    wamid = client.send_template_with_components(
                        to=r.wa_digits,
                        name=template,
                        language=language,
                        body_variables=[r.name],
                        header_video_id=media_id,
                    )
                    _dao.mark_sent(s, campaign=campaign, wa_digits=r.wa_digits, wamid=wamid)
                log.info(
                    "sent to=%s wamid=%s", _redact_phone(r.wa_digits), wamid
                )
                sent += 1
                send_ok = True
                break
            except MetaTransientError as e:
                if attempt == 2:
                    log.warning(
                        "transient error exhausted retries to=%s reason=%s",
                        _redact_phone(r.wa_digits),
                        str(e),
                    )
                    with get_session() as s:
                        _dao.mark_failed_local(
                            s,
                            campaign=campaign,
                            wa_digits=r.wa_digits,
                            reason=str(e),
                        )
                    failed += 1
                    break
                else:
                    backoff = retry_backoff * (attempt + 1)
                    log.info(
                        "transient error attempt=%d to=%s sleeping=%.1fs reason=%s",
                        attempt,
                        _redact_phone(r.wa_digits),
                        backoff,
                        str(e),
                    )
                    time.sleep(backoff)
            except MetaInvalidMessage as e:
                log.warning(
                    "invalid message to=%s reason=%s",
                    _redact_phone(r.wa_digits),
                    str(e),
                )
                with get_session() as s:
                    _dao.mark_failed_local(
                        s,
                        campaign=campaign,
                        wa_digits=r.wa_digits,
                        reason=str(e),
                    )
                failed += 1
                break

        time.sleep(spacing_seconds)

    log.info(
        "broadcast complete campaign=%s tier=%s "
        "attempted=%d sent=%d failed=%d skipped=%d",
        campaign,
        tier,
        attempted,
        sent,
        failed,
        skipped,
    )
    print(
        f"Broadcast summary: attempted={attempted} sent={sent} "
        f"failed={failed} skipped={skipped}"
    )
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Synchronous broadcast driver for munshi_welcome_video_v1.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--xlsx", required=True, help="Path to broadcast list xlsx.")
    p.add_argument("--campaign", required=True, help="Campaign identifier (used as DB key).")
    p.add_argument(
        "--tier",
        choices=["T1", "T2", "T3"],
        required=True,
        help="Tier to send to (matches 'Tier label' column prefix).",
    )
    p.add_argument(
        "--template",
        default=_DEFAULT_TEMPLATE,
        help="WhatsApp template name.",
    )
    p.add_argument(
        "--language",
        default=_DEFAULT_LANGUAGE,
        help="Template language code.",
    )
    p.add_argument(
        "--video-file",
        dest="video_file",
        default=None,
        help="Path to local video file to upload once before sending.",
    )
    p.add_argument(
        "--video-media-id",
        dest="video_media_id",
        default=None,
        help="Pre-uploaded Meta media_id for the video header (skips upload).",
    )
    p.add_argument(
        "--per-run-cap",
        dest="per_run_cap",
        type=int,
        default=_DEFAULT_PER_RUN_CAP,
        help="Max recipients per run (before daily-cap trim).",
    )
    p.add_argument(
        "--daily-cap",
        dest="daily_cap",
        type=int,
        default=_DEFAULT_DAILY_CAP,
        help="Max sends in the rolling 24h window.",
    )
    p.add_argument(
        "--spacing-seconds",
        dest="spacing_seconds",
        type=float,
        default=_DEFAULT_SPACING_SECONDS,
        help="Seconds to sleep between recipient sends.",
    )
    p.add_argument(
        "--retry-backoff",
        dest="retry_backoff",
        type=float,
        default=_DEFAULT_RETRY_BACKOFF,
        help="Backoff multiplier (sleep = retry_backoff * attempt_number).",
    )
    p.add_argument(
        "--dry-run",
        dest="dry_run",
        action="store_true",
        default=True,
        help="Log planned sends without actually sending (default ON).",
    )
    p.add_argument(
        "--yes",
        action="store_true",
        default=False,
        help="Disable dry-run and execute live sends.",
    )
    p.add_argument(
        "--max-fail-rate",
        dest="max_fail_rate",
        type=float,
        default=_DEFAULT_MAX_FAIL_RATE,
        help=(
            "Auto-pause guard: refuse to send if failed/attempted exceeds this "
            "fraction (requires --min-sample rows attempted). Default: 0.1."
        ),
    )
    p.add_argument(
        "--max-undeliverable",
        dest="max_undeliverable",
        type=float,
        default=_DEFAULT_MAX_UNDELIVERABLE,
        help=(
            "Auto-pause guard: refuse to send if undeliverable/attempted exceeds "
            "this fraction (requires --min-sample rows attempted). Default: 0.4."
        ),
    )
    p.add_argument(
        "--max-marketing-cap-rate",
        dest="max_marketing_cap_rate",
        type=float,
        default=_DEFAULT_MAX_MARKETING_CAP_RATE,
        help=(
            "Auto-pause guard: refuse to send if 131049 (marketing-cap) / "
            "attempted exceeds this fraction (requires --min-sample rows "
            "attempted). A sustained-high 131049 rate is Meta throttling the "
            "account for spam. Default: 0.3."
        ),
    )
    p.add_argument(
        "--min-sample",
        dest="min_sample",
        type=int,
        default=_DEFAULT_MIN_SAMPLE,
        help=(
            "Minimum number of attempted rows before the auto-pause guard "
            "activates. Guard is inactive for brand-new campaigns. Default: 50."
        ),
    )
    return p


def main(argv: list[str] | None = None) -> int:
    """CLI entry point. Returns exit code."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )
    parser = _build_parser()
    args = parser.parse_args(argv)

    # --yes overrides --dry-run
    if args.yes:
        args.dry_run = False

    # Validate: live send needs a video source
    if not args.dry_run and not args.video_file and not args.video_media_id:
        parser.error("live send requires --video-file or --video-media-id")

    return run(args)


if __name__ == "__main__":
    raise SystemExit(main())
